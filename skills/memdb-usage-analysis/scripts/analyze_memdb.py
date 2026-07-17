#!/usr/bin/env python3
"""Analyze Jstp Memdb showtables exports for one or many scenarios."""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from itertools import combinations
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


ANSI_RE = re.compile(r"\x1b\[[0-9;]*m")


@dataclass(frozen=True)
class Row:
    scenario: str
    table_name: str
    records: int
    alloc_mb: float
    field_count: int
    version: int
    disk_sync: str
    read_only: str


def read_text(path: Path) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gbk", "cp936"):
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return path.read_text(encoding="utf-8", errors="replace")


def parse_memdb_export(path: Path, scenario: str) -> list[Row]:
    rows: list[Row] = []
    for raw_line in read_text(path).splitlines():
        line = ANSI_RE.sub("", raw_line)
        if not line.startswith("|"):
            continue
        parts = [part.strip() for part in line.split("|")[1:-1]]
        if len(parts) != 7 or parts[0] == "内存表":
            continue
        try:
            rows.append(
                Row(
                    scenario=scenario,
                    table_name=parts[0],
                    records=int(parts[1]),
                    alloc_mb=float(parts[2]),
                    field_count=int(parts[3]),
                    version=int(parts[4]),
                    disk_sync=parts[5],
                    read_only=parts[6],
                )
            )
        except ValueError as exc:
            raise ValueError(f"Cannot parse data row in {path}: {raw_line}") from exc
    return rows


def read_csv_dicts(path: Path) -> list[dict[str, str]]:
    text = read_text(path)
    return list(csv.DictReader(text.splitlines()))


def write_csv(path: Path, rows: Iterable[dict[str, object]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def row_to_dict(row: Row) -> dict[str, object]:
    return {
        "Scenario": row.scenario,
        "TableName": row.table_name,
        "Records": row.records,
        "AllocMB": f"{row.alloc_mb:.2f}",
        "FieldCount": row.field_count,
        "Version": row.version,
        "DiskSync": row.disk_sync,
        "ReadOnly": row.read_only,
    }


def parse_mapping(values: list[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for value in values:
        if "=" not in value:
            raise ValueError(f"Expected KEY=VALUE mapping, got: {value}")
        key, mapped_value = value.split("=", 1)
        key = key.strip()
        mapped_value = mapped_value.strip()
        if not key or not mapped_value:
            raise ValueError(f"Invalid mapping: {value}")
        mapping[key] = mapped_value
    return mapping


def load_category_map(path: Path) -> tuple[dict[str, str], str]:
    rows = read_csv_dicts(path)
    if not rows:
        raise ValueError(f"Category config is empty: {path}")
    headers = rows[0].keys()
    if "TableName" not in headers:
        raise ValueError("Category config must contain TableName column.")
    if "分类" in headers:
        category_col = "分类"
    elif "Category" in headers:
        category_col = "Category"
    else:
        raise ValueError("Category config must contain 分类 or Category column.")

    mapping: dict[str, str] = {}
    duplicates: list[str] = []
    for row in rows:
        table_name = (row.get("TableName") or "").strip()
        category = (row.get(category_col) or "").strip()
        if not table_name:
            continue
        if table_name in mapping:
            duplicates.append(table_name)
        mapping[table_name] = category
    if duplicates:
        sample = ", ".join(sorted(set(duplicates))[:10])
        raise ValueError(f"Duplicate TableName values in category config: {sample}")
    return mapping, category_col


def safe_div(numerator: float, denominator: float) -> float | None:
    if denominator == 0:
        return None
    return numerator / denominator


def round_or_blank(value: float | None, digits: int) -> str:
    if value is None:
        return ""
    return f"{value:.{digits}f}"


def table_lookup(rows: list[Row]) -> dict[tuple[str, str], Row]:
    return {(row.scenario, row.table_name): row for row in rows}


def as_number(value: object) -> float | None:
    if value == "" or value is None:
        return None
    return float(value)


def excel_safe_name(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_]", "_", value)
    if not cleaned or cleaned[0].isdigit():
        cleaned = f"_{cleaned}"
    return cleaned[:200]


def add_table(ws, name: str, ref: str) -> None:
    table = Table(displayName=excel_safe_name(name), ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight10",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def style_report_sheet(ws) -> None:
    title_font = Font(name="等线", size=15, bold=True)
    header_font = Font(name="Carlito", size=11, color="FF000000")
    input_font = Font(name="等线", size=11, bold=True, color="FFFA7D00")
    input_fill = PatternFill("solid", fgColor="FFF2F2F2")
    center = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center
            if cell.value in {"内存占用分析", "平均账户内存占用分析", "计算节点容量计算"}:
                cell.font = title_font
            elif cell.row in {2} or cell.value in {
                "场景",
                "分类",
                "基准表",
                "条数",
                "总记录数",
                "总分配MB",
                "表数",
                "单条记录MB",
                "账户数",
                "单账户记录数",
                "单账户分配MB",
                "单账户总分配MB",
                "放大后记录数",
                "放大后单账户分配MB",
                "放大后单账户总分配MB",
                "参数类占用",
                "容纳客户数",
                "交易放大后容纳客户数",
            }:
                cell.font = header_font
    for coord in ("H22", "I22"):
        if ws[coord].value is not None:
            ws[coord].font = input_font
            ws[coord].fill = input_fill
    for coord in ("A35", "B35", "C35", "D35"):
        if ws[coord].value is not None:
            ws[coord].font = input_font
            ws[coord].fill = input_fill

    number_formats = {
        "D": "#,##0",
        "E": "#,##0",
        "F": "#,##0.00",
        "G": "#,##0",
        "H": "0.000000",
        "I": "0.000000",
        "J": "0.000000",
    }
    for col, fmt in number_formats.items():
        for cell in ws[col]:
            if cell.row > 1:
                cell.number_format = fmt

    widths = {
        "A": 22.8,
        "B": 13,
        "C": 20,
        "D": 18,
        "E": 18,
        "F": 18,
        "G": 23,
        "H": 27,
        "I": 27,
        "J": 24,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def write_excel_report(
    output_path: Path,
    category_summary: list[dict[str, object]],
    per_account_summary: list[dict[str, object]],
    scenario_names: list[str],
    parameter_category: str,
    compute_nodes: float,
    total_memory_mb: float,
    trade_amplification: float,
) -> None:
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "内存占用分析"
    headers1 = ["场景", "分类", "基准表", "条数", "总记录数", "总分配MB", "表数", "单条记录MB"]
    for col, header in enumerate(headers1, start=1):
        ws.cell(2, col, header)

    category_row_by_key: dict[tuple[str, str], int] = {}
    for offset, item in enumerate(category_summary, start=3):
        scenario = str(item["Scenario"])
        category = str(item["Category"])
        category_row_by_key[(scenario, category)] = offset
        ws.cell(offset, 1, scenario)
        ws.cell(offset, 2, category)
        ws.cell(offset, 3, item["BaselineTable"])
        baseline_records = item["BaselineRecordsAsTotalCount"]
        ws.cell(offset, 4, as_number(baseline_records))
        ws.cell(offset, 5, as_number(item["ActualRecordSumReference"]))
        ws.cell(offset, 6, as_number(item["TotalAllocMB"]))
        ws.cell(offset, 7, as_number(item["TableCountInScenario"]))
        if baseline_records != "":
            ws.cell(offset, 8, f"=F{offset}/D{offset}")

    summary_end = 2 + len(category_summary)
    if category_summary:
        add_table(ws, "表1", f"A2:H{summary_end}")

    section2_title = summary_end + 12
    ws.cell(section2_title, 1, "平均账户内存占用分析")
    ws.cell(section2_title, 8, "交易放大比例")
    ws.cell(section2_title, 9, trade_amplification)

    section2_header = section2_title + 1
    headers2 = [
        "场景",
        "分类",
        "账户数",
        "条数",
        "单账户记录数",
        "单账户分配MB",
        "单账户总分配MB",
        "放大后记录数",
        "放大后单账户分配MB",
        "放大后单账户总分配MB",
    ]
    for col, header in enumerate(headers2, start=1):
        ws.cell(section2_header, col, header)

    per_account_rows = [
        item for item in per_account_summary if str(item["Category"]) != parameter_category
    ]
    category_priority = {"账户类": 0, "交易类": 1, "资产类": 2, "持仓类": 2}
    per_account_rows.sort(
        key=lambda item: (
            scenario_names.index(str(item["Scenario"])) if str(item["Scenario"]) in scenario_names else 999,
            category_priority.get(str(item["Category"]), 50),
            str(item["Category"]),
        )
    )

    per_account_row_by_key: dict[tuple[str, str], int] = {}
    first_row_by_scenario: dict[str, int] = {}
    last_row_by_scenario: dict[str, int] = {}
    row_idx = section2_header + 1
    for item in per_account_rows:
        scenario = str(item["Scenario"])
        category = str(item["Category"])
        per_account_row_by_key[(scenario, category)] = row_idx
        first_row_by_scenario.setdefault(scenario, row_idx)
        last_row_by_scenario[scenario] = row_idx
        ws.cell(row_idx, 1, scenario)
        ws.cell(row_idx, 2, category)
        ws.cell(row_idx, 3, as_number(item["AccountCount"]))
        ws.cell(row_idx, 4, as_number(item["BaselineRecordsAsTotalCount"]))
        ws.cell(row_idx, 5, f"=D{row_idx}/C{row_idx}")
        summary_row = category_row_by_key[(scenario, category)]
        ws.cell(row_idx, 6, f"=F{summary_row}/C{row_idx}")
        if category == "账户类":
            ws.cell(row_idx, 8, f"=E{row_idx}")
            ws.cell(row_idx, 9, f"=F{row_idx}")
        else:
            ws.cell(row_idx, 8, f"=E{row_idx}*$I${section2_title}")
            ws.cell(row_idx, 9, f"=F{row_idx}*$I${section2_title}")
        row_idx += 1

    for scenario, first_row in first_row_by_scenario.items():
        last_row = last_row_by_scenario[scenario]
        ws.cell(first_row, 7, f"=SUM(F{first_row}:F{last_row})")
        ws.cell(first_row, 10, f"=SUM(I{first_row}:I{last_row})")

    section2_end = row_idx - 1
    if per_account_rows:
        add_table(ws, "表2", f"A{section2_header}:J{section2_end}")

    cap_title = section2_end + 5
    ws.cell(cap_title, 1, "计算节点容量计算")
    assumption_row = cap_title + 1
    ws.cell(assumption_row, 1, "计算节点数")
    ws.cell(assumption_row, 2, compute_nodes)
    ws.cell(assumption_row, 3, "总内存数据库大小")
    ws.cell(assumption_row, 4, total_memory_mb)

    cap_header = cap_title + 2
    for col, header in enumerate(["场景", "参数类占用", "容纳客户数", "交易放大后容纳客户数"], start=1):
        ws.cell(cap_header, col, header)

    cap_row = cap_header + 1
    for scenario in scenario_names:
        if scenario not in first_row_by_scenario:
            continue
        param_row = category_row_by_key.get((scenario, parameter_category))
        first_row = first_row_by_scenario[scenario]
        ws.cell(cap_row, 1, scenario)
        ws.cell(cap_row, 2, f"=$B${assumption_row}*F{param_row}" if param_row else 0)
        ws.cell(cap_row, 3, f"=($D${assumption_row}-B{cap_row})/G{first_row}")
        ws.cell(cap_row, 4, f"=($D${assumption_row}-B{cap_row})/J{first_row}")
        cap_row += 1

    cap_end = cap_row - 1
    if cap_end >= cap_header + 1:
        add_table(ws, "表3", f"A{cap_header}:D{cap_end}")

    style_report_sheet(ws)
    for row in range(3, summary_end + 1):
        ws.cell(row, 8).number_format = "0.000000"
    for row in range(section2_header + 1, section2_end + 1):
        for col in (5, 6, 7, 8, 9, 10):
            ws.cell(row, col).number_format = "0.000000"
    for row in range(cap_header + 1, cap_end + 1):
        for col in (2, 3, 4):
            ws.cell(row, col).number_format = "#,##0"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--scenario",
        action="append",
        required=True,
        help="Scenario mapping as name=path. Repeat for multiple scenarios.",
    )
    parser.add_argument("--category-config", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    parser.add_argument("--min-alloc-mb", type=float, default=10.0)
    parser.add_argument("--parameter-category", default="参数类")
    parser.add_argument("--account-baseline", default="node_fundacct")
    parser.add_argument("--excel-output", type=Path, help="Optional .xlsx output path. Defaults to output-dir/memdb_memory_usage_analysis.xlsx.")
    parser.add_argument("--compute-nodes", type=float, default=8.0, help="Compute node count used in capacity calculation.")
    parser.add_argument("--total-memory-mb", type=float, default=512.0 * 1024.0, help="Total Memdb capacity in MB used in capacity calculation.")
    parser.add_argument("--trade-amplification", type=float, default=1.5, help="Amplification factor for the capacity stress calculation.")
    parser.add_argument(
        "--compare-scenarios",
        action="store_true",
        help="Generate pairwise table-level differences. By default scenarios are processed independently.",
    )
    parser.add_argument(
        "--baseline",
        action="append",
        default=[],
        help="Category baseline mapping as category=table. Repeat as needed.",
    )
    args = parser.parse_args()

    scenario_paths = parse_mapping(args.scenario)
    baselines = {
        "交易类": "node_settdetail",
        "账户类": "node_fundacct",
        "资产类": "node_stkasset",
        "持仓类": "node_stkasset",
    }
    baselines.update(parse_mapping(args.baseline))

    output_dir: Path = args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    all_rows: list[Row] = []
    for scenario, raw_path in scenario_paths.items():
        path = Path(raw_path)
        if not path.exists():
            raise FileNotFoundError(path)
        parsed = parse_memdb_export(path, scenario)
        if not parsed:
            raise ValueError(f"No table rows parsed from {path}")
        all_rows.extend(parsed)

    filtered_rows = [row for row in all_rows if row.alloc_mb >= args.min_alloc_mb]
    category_map, category_col = load_category_map(args.category_config)
    retained_names = sorted({row.table_name for row in filtered_rows})

    presence_rows: list[dict[str, object]] = []
    for table_name in retained_names:
        out: dict[str, object] = {"TableName": table_name}
        for scenario in scenario_paths:
            out[f"In_{scenario}"] = any(
                row.scenario == scenario and row.table_name == table_name
                for row in filtered_rows
            )
        presence_rows.append(out)

    missing_names = [name for name in retained_names if name not in category_map or not category_map[name]]
    extra_names = sorted(set(category_map) - set(retained_names))

    write_csv(
        output_dir / "memdb_all_tables.csv",
        [row_to_dict(row) for row in all_rows],
        ["Scenario", "TableName", "Records", "AllocMB", "FieldCount", "Version", "DiskSync", "ReadOnly"],
    )
    threshold_label = ("%g" % args.min_alloc_mb).replace(".", "_")
    write_csv(
        output_dir / f"memdb_filtered_alloc_ge_{threshold_label}mb.csv",
        [row_to_dict(row) for row in filtered_rows],
        ["Scenario", "TableName", "Records", "AllocMB", "FieldCount", "Version", "DiskSync", "ReadOnly"],
    )
    presence_fields = ["TableName"] + [f"In_{scenario}" for scenario in scenario_paths]
    write_csv(output_dir / "memdb_filtered_dedup_table_presence.csv", presence_rows, presence_fields)

    if missing_names:
        missing_rows: list[dict[str, object]] = []
        for table_name in missing_names:
            out = {"TableName": table_name, category_col: ""}
            for scenario in scenario_paths:
                row = next(
                    (
                        item
                        for item in filtered_rows
                        if item.scenario == scenario and item.table_name == table_name
                    ),
                    None,
                )
                out[f"{scenario}_Records"] = row.records if row else ""
                out[f"{scenario}_AllocMB"] = f"{row.alloc_mb:.2f}" if row else ""
            missing_rows.append(out)
        missing_fields = ["TableName", category_col]
        for scenario in scenario_paths:
            missing_fields.extend([f"{scenario}_Records", f"{scenario}_AllocMB"])
        write_csv(output_dir / "missing_tables_need_classification.csv", missing_rows, missing_fields)
        print(
            json.dumps(
                {
                    "status": "missing_categories",
                    "message": "Retained tables are missing from the category config. Update the config and rerun.",
                    "missing_count": len(missing_names),
                    "missing_file": str(output_dir / "missing_tables_need_classification.csv"),
                },
                ensure_ascii=False,
                indent=2,
            )
        )
        return 2

    joined_rows: list[dict[str, object]] = []
    for row in filtered_rows:
        joined_rows.append(
            {
                **row_to_dict(row),
                "Category": category_map[row.table_name],
            }
        )
    write_csv(
        output_dir / "memdb_filtered_with_categories.csv",
        joined_rows,
        ["Scenario", "TableName", "Records", "AllocMB", "FieldCount", "Version", "DiskSync", "ReadOnly", "Category"],
    )

    categories = sorted({category_map[name] for name in retained_names})
    preferred_order = [args.parameter_category, "交易类", "账户类", "资产类", "持仓类"]
    ordered_categories = [cat for cat in preferred_order if cat in categories]
    ordered_categories.extend(cat for cat in categories if cat not in ordered_categories)

    lookup = table_lookup(filtered_rows)
    category_summary: list[dict[str, object]] = []
    per_account_summary: list[dict[str, object]] = []

    for scenario in scenario_paths:
        account_row = lookup.get((scenario, args.account_baseline))
        if account_row is None:
            raise ValueError(f"Missing account baseline {args.account_baseline} in scenario {scenario}")
        account_count = account_row.records
        scenario_rows = [row for row in filtered_rows if row.scenario == scenario]

        for category in ordered_categories:
            class_rows = [row for row in scenario_rows if category_map[row.table_name] == category]
            total_alloc_mb = sum(row.alloc_mb for row in class_rows)
            actual_record_sum = sum(row.records for row in class_rows)
            baseline_table = "" if category == args.parameter_category else baselines.get(category, "")
            baseline_records: int | None = None
            if baseline_table:
                baseline_row = lookup.get((scenario, baseline_table))
                if baseline_row is None:
                    raise ValueError(f"Missing baseline {baseline_table} for {scenario} / {category}")
                baseline_records = baseline_row.records

            avg_kb_per_record = (
                safe_div(total_alloc_mb * 1024.0, float(baseline_records))
                if baseline_records is not None
                else None
            )
            records_per_account = (
                safe_div(float(baseline_records), float(account_count))
                if baseline_records is not None
                else None
            )
            alloc_kb_per_account = safe_div(total_alloc_mb * 1024.0, float(account_count))

            category_summary.append(
                {
                    "Scenario": scenario,
                    "Category": category,
                    "BaselineTable": baseline_table,
                    "BaselineRecordsAsTotalCount": baseline_records if baseline_records is not None else "",
                    "ActualRecordSumReference": actual_record_sum,
                    "TotalAllocMB": f"{total_alloc_mb:.2f}",
                    "AvgKBPerBaselineRecord": round_or_blank(avg_kb_per_record, 4),
                    "TableCountInScenario": len(class_rows),
                }
            )
            per_account_summary.append(
                {
                    "Scenario": scenario,
                    "Category": category,
                    "AccountBaselineTable": args.account_baseline,
                    "AccountCount": account_count,
                    "BaselineRecordsAsTotalCount": baseline_records if baseline_records is not None else "",
                    "RecordsPerAccount": round_or_blank(records_per_account, 6),
                    "AllocKBPerAccount": round_or_blank(alloc_kb_per_account, 4),
                    "AllocMBPerAccount": round_or_blank(
                        None if alloc_kb_per_account is None else alloc_kb_per_account / 1024.0,
                        6,
                    ),
                }
            )

    write_csv(
        output_dir / "memdb_category_summary_by_scenario.csv",
        category_summary,
        [
            "Scenario",
            "Category",
            "BaselineTable",
            "BaselineRecordsAsTotalCount",
            "ActualRecordSumReference",
            "TotalAllocMB",
            "AvgKBPerBaselineRecord",
            "TableCountInScenario",
        ],
    )
    write_csv(
        output_dir / "memdb_category_per_account_by_scenario.csv",
        per_account_summary,
        [
            "Scenario",
            "Category",
            "AccountBaselineTable",
            "AccountCount",
            "BaselineRecordsAsTotalCount",
            "RecordsPerAccount",
            "AllocKBPerAccount",
            "AllocMBPerAccount",
        ],
    )

    excel_output = args.excel_output or (output_dir / "memdb_memory_usage_analysis.xlsx")
    write_excel_report(
        excel_output,
        category_summary,
        per_account_summary,
        list(scenario_paths.keys()),
        args.parameter_category,
        args.compute_nodes,
        args.total_memory_mb,
        args.trade_amplification,
    )

    validation_rows = []
    by_category: dict[str, int] = defaultdict(int)
    for name in retained_names:
        by_category[category_map[name]] += 1
    for category in ordered_categories:
        validation_rows.append(
            {
                "Category": category,
                "DedupRetainedTableCount": by_category[category],
                "BaselineTable": "" if category == args.parameter_category else baselines.get(category, ""),
            }
        )
    write_csv(
        output_dir / "memdb_category_validation_summary.csv",
        validation_rows,
        ["Category", "DedupRetainedTableCount", "BaselineTable"],
    )

    pairwise_count = 0
    if args.compare_scenarios and len(scenario_paths) >= 2:
        pair_rows: list[dict[str, object]] = []
        for scenario_a, scenario_b in combinations(scenario_paths.keys(), 2):
            pairwise_count += 1
            account_a = lookup[(scenario_a, args.account_baseline)].records
            account_b = lookup[(scenario_b, args.account_baseline)].records
            for table_name in retained_names:
                row_a = lookup.get((scenario_a, table_name))
                row_b = lookup.get((scenario_b, table_name))
                alloc_a = row_a.alloc_mb if row_a else 0.0
                alloc_b = row_b.alloc_mb if row_b else 0.0
                records_a = row_a.records if row_a else 0
                records_b = row_b.records if row_b else 0
                kb_a = safe_div(alloc_a * 1024.0, float(account_a)) or 0.0
                kb_b = safe_div(alloc_b * 1024.0, float(account_b)) or 0.0
                pair_rows.append(
                    {
                        "ScenarioA": scenario_a,
                        "ScenarioB": scenario_b,
                        "Category": category_map[table_name],
                        "TableName": table_name,
                        "ScenarioARecords": records_a,
                        "ScenarioAAllocMB": f"{alloc_a:.2f}",
                        "ScenarioAKBPerAccount": f"{kb_a:.4f}",
                        "ScenarioBRecords": records_b,
                        "ScenarioBAllocMB": f"{alloc_b:.2f}",
                        "ScenarioBKBPerAccount": f"{kb_b:.4f}",
                        "DiffRecords_BMinusA": records_b - records_a,
                        "DiffAllocMB_BMinusA": f"{alloc_b - alloc_a:.2f}",
                        "DiffKBPerAccount_BMinusA": f"{kb_b - kb_a:.4f}",
                    }
                )
        write_csv(
            output_dir / "memdb_table_diff_pairwise.csv",
            pair_rows,
            [
                "ScenarioA",
                "ScenarioB",
                "Category",
                "TableName",
                "ScenarioARecords",
                "ScenarioAAllocMB",
                "ScenarioAKBPerAccount",
                "ScenarioBRecords",
                "ScenarioBAllocMB",
                "ScenarioBKBPerAccount",
                "DiffRecords_BMinusA",
                "DiffAllocMB_BMinusA",
                "DiffKBPerAccount_BMinusA",
            ],
        )
    elif args.compare_scenarios and len(scenario_paths) < 2:
        print(
            json.dumps(
                {
                    "status": "comparison_skipped",
                    "message": "--compare-scenarios requires at least two scenarios; independent statistics were still generated.",
                },
                ensure_ascii=False,
                indent=2,
            )
        )

    result = {
        "status": "ok",
        "scenario_count": len(scenario_paths),
        "all_rows": len(all_rows),
        "filtered_rows": len(filtered_rows),
        "dedup_retained_tables": len(retained_names),
        "pairwise_comparisons": pairwise_count,
        "comparison_requested": bool(args.compare_scenarios),
        "extra_config_tables_ignored": len(extra_names),
        "output_dir": str(output_dir),
        "excel_output": str(excel_output),
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        raise SystemExit(1)
